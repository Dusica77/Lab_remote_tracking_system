from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime
import sqlite3
import qrcode
import json
import io
import base64
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
CORS(app)

# Initialize database
def init_db():
    conn = sqlite3.connect('lab_tracking.db')
    c = conn.cursor()
    
    # Create tables
    c.execute('''CREATE TABLE IF NOT EXISTS persons
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  email TEXT NOT NULL UNIQUE,
                  phone TEXT,
                  department TEXT,
                  registration_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS lab_records
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  person_id INTEGER,
                  lab_name TEXT NOT NULL,
                  entry_time TIMESTAMP,
                  exit_time TIMESTAMP,
                  FOREIGN KEY (person_id) REFERENCES persons (id))''')
    
    conn.commit()
    conn.close()

init_db()

@app.route('/api/register', methods=['POST'])
def register_person():
    data = request.json
    conn = sqlite3.connect('lab_tracking.db')
    c = conn.cursor()
    
    try:
        c.execute('''INSERT INTO persons (name, email, phone, department)
                     VALUES (?, ?, ?, ?)''',
                 (data['name'], data['email'], data['phone'], data['department']))
        
        person_id = c.lastrowid
        conn.commit()
        
        # Generate QR code data
        qr_data = {
            'id': person_id,
            'name': data['name'],
            'email': data['email']
        }
        
        # Generate QR code
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        buffered = io.BytesIO()
        img.save(buffered, format="PNG")
        qr_base64 = base64.b64encode(buffered.getvalue()).decode()
        
        return jsonify({
            'success': True,
            'person_id': person_id,
            'qr_code': qr_base64,
            'message': 'Person registered successfully'
        })
        
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': 'Email already exists'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/api/scan', methods=['POST'])
def scan_qr():
    data = request.json
    qr_content = data.get('qr_content')
    lab_name = data.get('lab_name', 'Main Lab')
    
    try:
        person_data = json.loads(qr_content)
        person_id = person_data['id']
        
        conn = sqlite3.connect('lab_tracking.db')
        c = conn.cursor()
        
        # Check if person has an open session
        c.execute('''SELECT id FROM lab_records 
                     WHERE person_id = ? AND exit_time IS NULL''', (person_id,))
        open_session = c.fetchone()
        
        if open_session:
            # Close the session (exit)
            c.execute('''UPDATE lab_records 
                         SET exit_time = ?
                         WHERE id = ?''', (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), open_session[0]))
            action = 'exit'
        else:
            # Create new session (entry)
            c.execute('''INSERT INTO lab_records (person_id, lab_name, entry_time)
                         VALUES (?, ?, ?)''', (person_id, lab_name, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            action = 'entry'
        
        conn.commit()
        
        # Get person details
        c.execute('SELECT name, email FROM persons WHERE id = ?', (person_id,))
        person = c.fetchone()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'action': action,
            'person': {
                'name': person[0],
                'email': person[1]
            },
            'lab_name': lab_name,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/records', methods=['GET'])
def get_records():
    conn = sqlite3.connect('lab_tracking.db')
    c = conn.cursor()
    
    c.execute('''SELECT lr.id, p.name, p.email, lr.lab_name, 
                 lr.entry_time, lr.exit_time 
                 FROM lab_records lr
                 JOIN persons p ON lr.person_id = p.id
                 ORDER BY lr.entry_time DESC''')
    
    records = []
    for row in c.fetchall():
        records.append({
            'id': row[0],
            'name': row[1],
            'email': row[2],
            'lab_name': row[3],
            'entry_time': row[4],
            'exit_time': row[5]
        })
    
    conn.close()
    return jsonify(records)

@app.route('/api/current_lab_status', methods=['GET'])
def get_current_lab_status():
    conn = sqlite3.connect('lab_tracking.db')
    c = conn.cursor()
    
    # Get labs with current occupants
    c.execute('''SELECT lr.lab_name, p.name, p.email, lr.entry_time
                 FROM lab_records lr
                 JOIN persons p ON lr.person_id = p.id
                 WHERE lr.exit_time IS NULL
                 ORDER BY lr.entry_time DESC''')
    
    current_occupants = []
    for row in c.fetchall():
        current_occupants.append({
            'lab_name': row[0],
            'name': row[1],
            'email': row[2],
            'entry_time': row[3]
        })
    
    # Get last exited lab for each person
    c.execute('''SELECT lr.lab_name, p.name, MAX(lr.exit_time) as last_exit
                 FROM lab_records lr
                 JOIN persons p ON lr.person_id = p.id
                 WHERE lr.exit_time IS NOT NULL
                 GROUP BY p.id
                 ORDER BY last_exit DESC''')
    
    last_exits = []
    for row in c.fetchall():
        last_exits.append({
            'lab_name': row[0],
            'name': row[1],
            'last_exit': row[2]
        })
    
    conn.close()
    
    return jsonify({
        'current_occupants': current_occupants,
        'last_exits': last_exits
    })

@app.route('/api/person/<int:person_id>', methods=['GET'])
def get_person_by_id(person_id):
    conn = sqlite3.connect('lab_tracking.db')
    c = conn.cursor()
    
    try:
        # Get person details from persons table
        c.execute('SELECT id, name, email, phone, department FROM persons WHERE id = ?', (person_id,))
        person = c.fetchone()
        
        if person:
            return jsonify({
                'success': True,
                'person': {
                    'id': person[0],
                    'name': person[1],
                    'email': person[2],
                    'phone': person[3],
                    'department': person[4]
                }
            })
        else:
            return jsonify({
                'success': False,
                'message': f'No person found with ID: {person_id}'
            })
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/api/records/<int:record_id>', methods=['DELETE'])
def delete_record(record_id):
    conn = sqlite3.connect('lab_tracking.db')
    c = conn.cursor()
    try:
        c.execute('DELETE FROM lab_records WHERE id = ?', (record_id,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Record deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/api/records', methods=['DELETE'])
def delete_all_records():
    conn = sqlite3.connect('lab_tracking.db')
    c = conn.cursor()
    try:
        c.execute('DELETE FROM lab_records')
        conn.commit()
        return jsonify({'success': True, 'message': 'All records deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

# Excel Export Routes
@app.route('/api/export/excel', methods=['GET'])
def export_to_excel():
    conn = sqlite3.connect('lab_tracking.db')
    
    try:
        # Get all records with person details
        query = '''
        SELECT 
            p.id as person_id,
            p.name,
            p.email,
            p.phone,
            p.department,
            lr.lab_name,
            lr.entry_time,
            lr.exit_time,
            CASE 
                WHEN lr.exit_time IS NULL THEN 'IN LAB'
                ELSE 'LEFT LAB'
            END as status
        FROM lab_records lr
        JOIN persons p ON lr.person_id = p.id
        ORDER BY lr.entry_time DESC
        '''
        
        df = pd.read_sql_query(query, conn)
        
        # Create Excel file in memory
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main records sheet
            df.to_excel(writer, sheet_name='Lab Records', index=False)
            
            # Summary sheet
            summary_data = {
                'Metric': ['Total Records', 'Current Lab Occupants', 'Unique Persons', 'Date Generated'],
                'Value': [
                    len(df),
                    len(df[df['status'] == 'IN LAB']),
                    df['person_id'].nunique(),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Get workbook and worksheets for styling
            workbook = writer.book
            records_sheet = workbook['Lab Records']
            summary_sheet = workbook['Summary']
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for sheet in [records_sheet, summary_sheet]:
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for sheet in [records_sheet, summary_sheet]:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'lab_records_export_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

@app.route('/api/export/current_status', methods=['GET'])
def export_current_status():
    conn = sqlite3.connect('lab_tracking.db')
    
    try:
        # Get current lab occupants
        query = '''
        SELECT 
            p.name,
            p.email,
            p.department,
            p.phone,
            lr.lab_name,
            lr.entry_time
        FROM lab_records lr
        JOIN persons p ON lr.person_id = p.id
        WHERE lr.exit_time IS NULL
        ORDER BY lr.entry_time DESC
        '''
        
        df = pd.read_sql_query(query, conn)
        
        # Create Excel file
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Current Lab Status', index=False)
            
            # Style the workbook
            workbook = writer.book
            sheet = workbook['Current Lab Status']
            
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust columns
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'current_lab_status_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})
    finally:
        conn.close()

if __name__ == '__main__':
    app.run(debug=True, port=5000, host='0.0.0.0')